import openpyxl

class ExcelUtils:
    """
    Utility class for interacting with Excel files using openpyxl.
    This class provides methods to read and write data in an Excel sheet.
    """

    @staticmethod
    def get_row_count(file_path, sheet_name):
        """
        Get the total number of rows in the given Excel sheet.
        :param file_path: Path to the Excel file.
        :param sheet_name: Name of the sheet to read from.
        :return: Total row count.
        """
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        return sheet.max_row

    @staticmethod
    def get_column_count(file_path, sheet_name):
        """
        Get the total number of columns in the given Excel sheet.
        :param file_path: Path to the Excel file.
        :param sheet_name: Name of the sheet to read from.
        :return: Total column count.
        """
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        return sheet.max_column

    @staticmethod
    def read_data(file_path, sheet_name, row, column):
        """
        Read data from a specific cell in an Excel sheet.
        :param file_path: Path to the Excel file.
        :param sheet_name: Name of the sheet to read from.
        :param row: Row number (starting from 1).
        :param column: Column number (starting from 1).
        :return: Data from the specified cell.
        """
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        return sheet.cell(row=row, column=column).value

    @staticmethod
    def write_data(file_path, sheet_name, row, column, data):
        """
        Write data to a specific cell in an Excel sheet.
        :param file_path: Path to the Excel file.
        :param sheet_name: Name of the sheet to write to.
        :param row: Row number (starting from 1).
        :param column: Column number (starting from 1).
        :param data: Data to be written.
        """
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        sheet.cell(row=row, column=column).value = data
        workbook.save(file_path)
        workbook.close()
